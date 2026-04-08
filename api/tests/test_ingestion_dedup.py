"""
Tests for Phase 18: content-hash dedup + MIME-first registry routing.
"""
from __future__ import annotations

import hashlib
from pathlib import Path
from unittest.mock import MagicMock, patch
from uuid import uuid4

import pytest

from app.db.models import Upload, User
from app.db.session import SessionLocal
from app.services.ingestion_service import IngestionService, IngestionError
from app.core.config import get_settings


def _make_service(db):
    minio = MagicMock()
    minio.put_object = MagicMock()
    settings = get_settings()
    return IngestionService(db=db, settings=settings, minio_client=minio)


def _create_user(db, user_id: str) -> User:
    """Create a minimal user row required by FK constraints."""
    user = User(
        id=user_id,
        email=f"{user_id[:8]}@test.invalid",
        password_hash="x",
        is_admin=False,
    )
    db.add(user)
    db.flush()
    return user


# ---------------------------------------------------------------------------
# Dedup tests
# ---------------------------------------------------------------------------

def test_first_upload_succeeds(reset_db):
    db = SessionLocal()
    try:
        user_id = str(uuid4())
        _create_user(db, user_id)
        db.commit()
        svc = _make_service(db)
        file_bytes = b"dummy bank statement content unique 123"
        result = svc.upload_file(
            user_id=user_id,
            filename="statement.csv",
            content_type="text/csv",
            file_bytes=file_bytes,
        )
        assert "upload_id" in result
        upload = db.query(Upload).filter_by(user_id=user_id).first()
        assert upload is not None
        assert upload.content_hash == hashlib.sha256(file_bytes).hexdigest()
    finally:
        db.close()


def test_duplicate_rejected(reset_db):
    db = SessionLocal()
    try:
        user_id = str(uuid4())
        _create_user(db, user_id)
        db.commit()
        svc = _make_service(db)
        file_bytes = b"duplicate file content abc"
        svc.upload_file(
            user_id=user_id,
            filename="file.pdf",
            content_type="application/pdf",
            file_bytes=file_bytes,
        )
        with pytest.raises(IngestionError) as exc_info:
            svc.upload_file(
                user_id=user_id,
                filename="file_copy.pdf",
                content_type="application/pdf",
                file_bytes=file_bytes,
            )
        assert exc_info.value.code == "duplicate_file"
        assert exc_info.value.status_code == 409
    finally:
        db.close()


def test_same_bytes_different_user_not_duplicate(reset_db):
    db = SessionLocal()
    try:
        user_id_a = str(uuid4())
        user_id_b = str(uuid4())
        _create_user(db, user_id_a)
        _create_user(db, user_id_b)
        db.commit()
        svc = _make_service(db)
        file_bytes = b"same content for different users"
        svc.upload_file(
            user_id=user_id_a,
            filename="file.pdf",
            content_type="application/pdf",
            file_bytes=file_bytes,
        )
        # Different user → should succeed
        result = svc.upload_file(
            user_id=user_id_b,
            filename="file.pdf",
            content_type="application/pdf",
            file_bytes=file_bytes,
        )
        assert "upload_id" in result
    finally:
        db.close()


# ---------------------------------------------------------------------------
# MIME-first registry routing
# ---------------------------------------------------------------------------

def test_registry_loads_new_modules():
    """All 4 new builtin modules must load without error."""
    from app.ingestion.registry import ModuleRegistry
    reg = ModuleRegistry()
    names = {m.file_path.stem for m in reg.modules}
    for expected in ("payslip_it", "utility_bill_it", "invoice_it", "receipt_it"):
        assert expected in names, f"{expected} not loaded by registry"


def test_xlsx_fingerprint_uses_cell_text(tmp_path):
    """XLSX module should match when fingerprint keywords appear in cell text."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Conto Corrente"
    ws["A2"] = "Data_Operazione"
    ws["A3"] = "Entrate"
    ws["A4"] = "Uscite"
    ws["A5"] = "Descrizione_Completa"
    ws["A6"] = "Risultati Ricerca"
    xlsx_path = tmp_path / "fineco_test.xlsx"
    wb.save(str(xlsx_path))

    from app.ingestion.registry import ModuleRegistry
    reg = ModuleRegistry()
    mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    match = reg.match(xlsx_path, mime_type=mime)
    assert match is not None, "Expected fineco_it module to match the XLSX fixture"
    assert "fineco" in match.file_path.stem.lower()
